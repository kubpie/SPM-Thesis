# -*- coding: utf-8 -*-

"""
Created on Fri Apr 10 18:26:12 2020

@author: kubap
"""
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def find_extrema(ssp):
    
    a = np.array(ssp)
    minima = np.where(np.r_[True, a[1:] < a[:-1]] & np.r_[a[:-1] < a[1:], True])
    minima = minima[0][:].tolist()

    maxima = np.where(np.r_[True, a[1:] > a[:-1]] & np.r_[a[:-1] > a[1:], True])
    maxima = maxima[0][:].tolist()
    return minima, maxima

def SSPGrad(SSP_Input, path, save = False):
     
    depth = SSP_Input['DEPTH'].values.tolist()
    SSP_input = SSP_Input.iloc[:,1:]
  
    #Calculate gradient at each point of the depth-grid
    SSP_Grad = np.zeros(SSP_input.shape) #discrete gradient
    for i, ssp in enumerate(SSP_input):   
        #1. Calculate gradient at each point of the depth-grid
        for row in range(1,len(SSP_input)):
            grad = (SSP_input[ssp][row]-SSP_input[ssp][row-1])/(depth[row]-depth[row-1])
            SSP_Grad[row,i] = grad
    SSP_Grad = pd.DataFrame(SSP_Grad, index=np.arange(len(SSP_input)), columns = SSP_input.columns)       
   
    if save == True:        
        SSP_Grad = pd.concat([pd.DataFrame(depth, columns = ["depth"]), SSP_Grad], axis = 1)
        append_df_to_excel(path + 'env.xlsx', df = SSP_Grad, sheet_name='SSP_GRAD', startrow=0,  truncate_sheet=True, index = False)    
        #SSP_Grad.to_csv(path+"SSP_Grad.csv", index = False)
    
    return SSP_Grad

def SSPStat(SSP_Input, path, plot = False, save = False):
    
    SSP_Grad = SSPGrad(SSP_Input, path, save = False)

    depth = SSP_Input['DEPTH'].values.tolist()
    SSP_input = SSP_Input.iloc[:,1:]
    
    stat_SSP = ['mean_SSP','stdev_SSP', 'mean_grad','stdev_grad']
    colindex = pd.MultiIndex.from_product([SSP_input, stat_SSP])
    SSP_Stat = pd.DataFrame(np.zeros([np.size(SSP_input,0),np.size(SSP_input,1)*4]),columns = colindex) 
    for ssp in SSP_input.columns:
        for row in range(len(SSP_input)):
            
            SSP_Stat.loc[row,(ssp,'mean_SSP')] = np.mean(SSP_input[ssp].iloc[:row+1])
            SSP_Stat.loc[row,(ssp,'stdev_SSP')] = np.std(SSP_input[ssp].iloc[:row+1])

            SSP_Stat.loc[row,(ssp,'mean_grad')] = np.mean(SSP_Grad[ssp].iloc[:row+1])
            SSP_Stat.loc[row,(ssp,'stdev_grad')] = np.std(SSP_Grad[ssp].iloc[:row+1])
    if plot == True:
            fig, axes = plt.subplots(nrows = 3, ncols = 8, figsize = (10,5), sharey = True, sharex = True)
            axes = axes.flat    
            
            for i, ssp in enumerate(SSP_input):
                lower_bound = SSP_Stat.loc[:,(ssp,'mean_SSP')]-SSP_Stat.loc[:,(ssp,'stdev_SSP')]
                upper_bound = SSP_Stat.loc[:,(ssp,'mean_SSP')]+SSP_Stat.loc[:,(ssp,'stdev_SSP')]
                axes[i].plot(depth, np.array(SSP_Stat.loc[:,(ssp,'mean_SSP')]), linewidth = 1, label = 'Mean SSP' )
                axes[i].fill_between(depth, lower_bound, upper_bound, facecolor='lightblue', label = 'Std. Dev. SSP')
                axes[i].set_title("{}. {}".format(i, ssp))
            
            #TODO: Fix labels here and SSP plot
            #axes.set_xlabel("Depth")
            #axes.set_ylabel("SSP")
            handles, labels = axes[0].get_legend_handles_labels()
            fig.legend(handles, labels, ncol = 2, loc='upper center')
            plt.ylim([1450,1550])
            plt.show()
            
            fig, axes = plt.subplots(nrows = 3, ncols = 8, figsize = (10,5), sharey = True, sharex = True)
            axes = axes.flat    
            
            for i, ssp in enumerate(SSP_Grad):
                lower_bound = SSP_Stat.loc[:,(ssp,'mean_grad')]-SSP_Stat.loc[:,(ssp,'stdev_grad')]
                upper_bound = SSP_Stat.loc[:,(ssp,'mean_grad')]+SSP_Stat.loc[:,(ssp,'stdev_grad')]
                axes[i].plot(depth, np.array(SSP_Stat.loc[:,(ssp,'mean_grad')]), linewidth = 1, color = 'red', label = 'Mean SSP Gradient' )
                axes[i].fill_between(depth, lower_bound, upper_bound, facecolor='lightcoral', label = 'Std. Dev. Grad SSP')
                axes[i].set_title("{}. {}".format(i, ssp))
            
            #TODO: Fix labels here and SSP plot
            #axes.set_xlabel("Depth")
            #axes.set_ylabel("SSP")
            handles, labels = axes[0].get_legend_handles_labels()
            fig.legend(handles, labels, ncol = 2, loc='upper center')
            plt.ylim([-0.6,0.2])
            plt.show()   
            
    if save == True:
        append_df_to_excel(path + 'env.xlsx', df = SSP_Stat, sheet_name='SSP_STAT', startrow = 0, truncate_sheet=True, index = True)

    return SSP_Stat

def SSPId(SSP_Input, path, plot = False, save = False):   
    
    #1. Calculate gradient at each point of the depth-grid
    SSP_Grad = SSPGrad(SSP_Input, path, save = False)
    
    depth = SSP_Input['DEPTH'].values.tolist()
    SSP_input = SSP_Input.iloc[:,1:]
    max_depth = [50, 150, 250, 350, 450, 600, 750, 900, 1050, 1200, 1500]
    
    ###TODO: MEAN GRADS CALCULATIONS AND FEATURES PROCESSING FOR GRAKN FORMAT ###
    
    allvalues = []
    for md in max_depth:
            
        #Cut Input Data SSP max_depth
        SSP_table = SSP_input.iloc[0:depth.index(md)+1,:]
        
        #Deep Channel props
        DC_axis = np.zeros([SSP_table.shape[1],2]).astype(int) #deep channel axis
        DC_limits = np.ones([SSP_table.shape[1],3]).astype(int)*(-1) #DC critical depth  
        DC_grad = np.ones([SSP_table.shape[1],3])*(-1)
        
        #Sonic Layer props
        SL_depth = np.zeros([SSP_table.shape[1],2]).astype(int) #sonic layer depth
        SLD_grad = np.ones([SSP_table.shape[1],2])*(-1)
                
        for i, ssp in enumerate(SSP_table):
            ### SSP IDENTIFICATION ### 
            #2. Calculate Sonic Layer Depth
            [minima, maxima] = find_extrema(SSP_table[ssp]) #find minima & maxima of the curve
            first_nonzero_max = next((mx for mx in maxima if 0 < mx < 35), 999)
            first_nonzero_min = next((mn for mn in minima if 0 < mn < 35), 999)
            
            if maxima[0] == len(SSP_table)-1:
                # half-channel    
                sldi = maxima[0]
                SL_depth[i,0] = i #index
                SL_depth[i,1] = sldi #depth
                dci = 0
            else:
                if first_nonzero_max < first_nonzero_min:
                    # SD without channels before
                    sldi = first_nonzero_max         
                    SL_depth[i,0] = i #index             
                    SL_depth[i,1] = sldi #depth
                    #SL_depth[i,2] = SSP_table[ssp].iloc[sldi] #ssp
                else:
                    sldi = 0
                    SL_depth[i,:] = -1
                    
            #3. Calculate Deep Channel axis and width
                dci = np.argmin(SSP_table[ssp][sldi:]) #gives index of SSP SSP_table[ssp], looking for minimum BELOW SLD!
                dci = dci + sldi
    
            if dci == 0 or dci == (len(SSP_table)-1):
                # DCax at 0 or end = linear SSP, no DC
                DC_axis[i,:] = -1
            else:
                #DC limited by the SSP maximum (top or bottom)
                closest_max = maxima[np.searchsorted(maxima,dci,side='right')-1] #where to put dci such that the order is maintained
                next_max_below = maxima[maxima.index(closest_max)+1] #next max BELOW DC axis
                #next_max_above = maxima[maxima.index(closest_max)-1]
                
                if SSP_table[ssp].iloc[closest_max] <= SSP_table[ssp].iloc[next_max_below]: #SSP_table[ssp].iloc[-1]:
                #DC limited from above
                    cdi = closest_max
                    ssp_cmax = SSP_table[ssp].iloc[cdi]      
                    ssp_conj = np.searchsorted(SSP_table[ssp].iloc[dci:],ssp_cmax,side='right')
                    ssp_conj = ssp_conj + dci
                else:
                    #DC limited from below
                    midmax = [m for m in maxima if m != 35]     
                    if any(midmax > dci):
                        #DC limited by maximum above bottom
                        cdi = [m for m, val in enumerate(midmax > dci) if val]
                        cdi = maxima[cdi[-1]]
                    else:
                        #DC limited by bottom
                        cdi = len(SSP_table)-1
                        
                    ssp_cmax = SSP_table[ssp].iloc[cdi]
                    
                    # different top boundary depending on condition: sld, local max, or sea surface
                    if sldi > 0:
                        reverse = SSP_table[ssp].iloc[dci-1:sldi-1:-1]
                    if closest_max > 0:
                        reverse = SSP_table[ssp].iloc[dci-1:closest_max-1:-1]
                    else:
                        reverse = SSP_table[ssp].iloc[dci-1::-1]
                        
                    reverse = reverse.reset_index(drop = True)
                    ssp_conj_rev = np.searchsorted(reverse,ssp_cmax,side='right')
                    if ssp_conj_rev == len(reverse):
                        ssp_conj_rev = ssp_conj_rev-1
                        #end of range correction
                    ssp_conj = SSP_table[ssp][SSP_table[ssp] == reverse.iloc[ssp_conj_rev]].index[0]
                    
                # Deep sound channel (SOFAR) parameters 
                DC_axis[i,0] = i
                DC_axis[i,1] = dci
                
                DC_limits[i,0] = i
                DC_limits[i,1] = cdi
                DC_limits[i,2] = ssp_conj
        
                ### END OF SSP IDENTIFICATION  ###
        
        # In depth loop
        #4. (Optional) Plot SSP with props for each depth max   
        if plot == True:
            
            fig, axes = plt.subplots(nrows = 3, ncols = 8, figsize = (15,20), sharey = True)
            axes = axes.flat
            axes[0].invert_yaxis()
            for i, ssp in enumerate(SSP_table):
                axes[i].plot(np.array(SSP_table[ssp]), depth[0:depth.index(md)+1], linewidth = 2, label = 'Sound Speed Profile' )
                axes[i].set_title("{}. {}".format(i, ssp))
                for j in range(len(SL_depth)):
                    if i == SL_depth[j,0] and SL_depth[j,1] != -1:    
                        axes[i].axhline(y = depth[SL_depth[j,1]], linestyle = '-', color = 'red', linewidth = 1, label = 'Sonic Layer Depth')
            
                for k in range(len(DC_axis)):
                    if i == DC_axis[k,0] and DC_axis[k,1] != -1:
                        axes[i].axhline(y = depth[DC_axis[k,1]], linestyle = '-', color='g', linewidth = 1, label = 'DC axis')
                        axes[i].axhline(y = depth[DC_limits[k,1]], linestyle = '-.', color='g', linewidth = 1, label = 'Critical Depth')
                        axes[i].axhline(y = depth[DC_limits[k,2]], linestyle = '-.', color='g', linewidth = 1, label = 'Conjugate Depth')
                        axes[i].text(SSP_table[ssp].iloc[DC_limits[k,1]]+0.1, depth[DC_limits[k,1]] , 'A', size=10)
                        axes[i].text(SSP_table[ssp].iloc[DC_limits[k,2]]+0.1, depth[DC_limits[k,2]] , 'A\'', size=10)
            
            handles, labels = axes[9].get_legend_handles_labels()
            fig.legend(handles, labels, ncol = 5, loc='upper center')
            
            plt.show()
                
        # In depth loop
        #5. Calculate average gradients above SLD and above\below DC-axis
        for i in range(len(SL_depth)):
            if all(SL_depth[i,:] != -1):
                SLD_grad[i,0] = SL_depth[i,0] #SSP idx
                SLD_grad[i,1] = np.mean(SSP_Grad.iloc[0:SL_depth[i,1]+1, i]) #mean grad above sld
        for j in range(len(DC_axis)):
            if all(DC_axis[j,:] != -1):
                DC_grad[j,0] = DC_axis[j,0] #idx
                DC_grad[j,1] = np.mean(SSP_Grad.iloc[min(DC_limits[j,1],DC_limits[j,2])+1:DC_axis[j,1]+1, j]) #grad above axis to top limit
                DC_grad[j,2] = np.mean(SSP_Grad.iloc[DC_axis[j,1]+1:max(DC_limits[j,1],DC_limits[j,2])+1, j]) #grad below axis to bottom limit
                
                #Switch DC_limits in place such that DC_limits[j,1] = 'DC_top'
                # is always smaller than  DC_limits[j,2] = 'DC_bott'
                DC_limits[j,1], DC_limits[j,2] = min(DC_limits[j,1], DC_limits[j,2]), max(DC_limits[j,1], DC_limits[j,2]) #'DC_top'
                
        #6. Turn indices to depths
        def DepthIdx(arr):
            for col in range(np.size(arr,1)):
                for row in range(np.size(arr,0)):
                    if arr[row,col] != -1:
                        arr[row,col] = depth[arr[row,col]]
            return arr
        #7. Collect data and append to list of arrays 'allvalues'
        dims = np.column_stack([SL_depth[:,1], DC_axis[:,1], DC_limits[:,1:3]]) 
        dims = DepthIdx(dims)
        
        grads = np.column_stack([SLD_grad[:,1], DC_grad[:,1:3]])
        
        values = np.column_stack([np.arange(0,24),
                                  np.ones(np.size(SSP_table,1))*md,
                                  dims[:,0], grads[:,0], 
                                  dims[:,1:], grads[:,1:]])
        
        """    
        # more useful to keep them
        #8. Drop rows with no identified features
        nan_idx = []
        for row in range(len(values)):
             test_nan = (values[row,2] == values[row,3:]).all(0)
             if test_nan == True:
                 nan_idx.append(row)
        values = np.delete(values, nan_idx, axis = 0)
        """
        
        allvalues.append(values)

    allvalues = np.concatenate(allvalues, axis=0)   
    # Replace -1 with NaNs (skip for XGB application)
    allvalues = np.where(allvalues==-1, None, allvalues)
    # Replace SSP idx with SSP names
    allnames = np.array([SSP_table.columns.tolist()[name] for name in allvalues[:,0].astype(int)],dtype = str)
    
    SSP_Prop = pd.concat([pd.DataFrame(allnames), pd.DataFrame(allvalues[:,1:])], axis = 1)
    SSP_Prop.columns = ['SSP','dmax', 'SLD_depth',  'SLD_avgrad', 'DC_axis', 
                            'DC_top', 'DC_bot', 'DC_avgrad_top', 'DC_avgrad_bot']
    # Sort by SSP name (alphabetical order) and ascending dmax
    SSP_Prop = SSP_Prop.sort_values(['SSP','dmax'], ascending = ['True', 'True'])
         
    #Create a DataFrame with OUTPUT lookup table, then export to .csv in the DATA folder
    if save == True:
        append_df_to_excel(path + 'env.xlsx', df = SSP_Prop, sheet_name='SSP_PROP', startrow = 0, truncate_sheet=True, index = False)

        #SSP_Prop.to_csv(path + 'SSP_prop.csv', index = False)
    return SSP_Prop
    ### End of SSP Identification


import os
path = os.getcwd()+'\data\\'
SSP_Input = pd.read_excel(path+"env.xlsx", sheet_name = "SSP")
#SSP_Grad = SSPGrad(SSP_Input, path, save = False)
#SSP_Stat = SSPStat(SSP_Input, path, plot = False, save = False)
#SSP_Prop = SSPId(SSP_Input, path, plot = True, save = False)

#######################################################################################
# TODO:  REFINE SSP APPROXIMATION -> REDUCE THE NR OF ATTRIBUTES AS MUCH AS POSSIBLE
 #1. Evaluate the error of downsampling SSP numerically
 #2. Alternatively: numpy.polyfit with deg < len(max_depth) = 12, 
 #   and find best approximation for each SSP cut
 
# Plot to see the effect of subsampling SSP for Grakn input
plot_sampling = False
if plot_sampling == True:
    fig, axes = plt.subplots(nrows = 3, ncols = 8, figsize = (15,20), sharey = True)
    axes = axes.flat
    axes[0].invert_yaxis()
    
    # This downsampling is used now in GRAKN. 0 is skipped in favor of first point at 10m
    # because grad at 0 is 0, which breaks the SSP-vec 'triplet'
    
    max_depth = [0, 50, 150, 250, 350, 450, 600, 750, 900, 1050, 1200, 1500]
    depth = SSP_Input.iloc[:,0]

    for i, ssp in enumerate(SSP_Input.iloc[:,1:]):
        axes[i].plot(SSP_Input.iloc[:,1:][ssp], depth, linewidth = 2, label = 'Sound Speed Profile' )
        axes[i].plot(SSP_Input.iloc[:,1:][ssp][depth.isin(max_depth)], depth[depth.isin(max_depth)], linewidth = 1, label = 'Subsampled Sound Speed Profile')
        axes[i].set_title("{}. {}".format(i, ssp))

#from numpy.polynomial import Polynomial as polyfit
import numpy.polynomial.polynomial as poly
from sklearn.preprocessing import normalize

def PolyfitSSP(SSP_Input):
    depth = SSP_Input.iloc[:,0]
    ssp_input = SSP_Input.iloc[:,1:]
    resid = []
    coeff = []
    rank = 10
    for ssp in ssp_input.columns:
        y = np.array(ssp_input[ssp]).astype(float)
        x = np.array(depth).astype(float)
        # TODO: Find out how to unbias the prediction        
        for deg in range(1,1+rank):
            c, [r, _, _, _] = poly.polyfit(x, y, deg, full = True)
            coeff.append(c)
            resid.append(r)
    allres = []  
    best = []
    for it in range(len(SSP_Input.columns)-1):
        allres.append([coeff[it:it+rank],resid[it:it+rank]])
        best_it = np.argmin(resid[it:it+rank])
        best_r = min(resid[it:it+rank])
        best_c = coeff[it:it+rank][best_it]

        best.append([best_it, best_r, best_c])
    
    return best, allres

best, allres = PolyfitSSP(SSP_Input)
deg = range(1,11)


z = np.array(SSP_Input.iloc[:,0]).astype(float)
znew = np.linspace(z[0], z[-1], num=len(z)*5)

#plt.plot(xnew,ffit,x,y)


fig, axes = plt.subplots(nrows = 3, ncols = 8, figsize = (15,20), sharey = True)
axes = axes.flat
axes[0].invert_yaxis()
for i, ssp in enumerate(SSP_Input.iloc[:,1:]):
    coeff = best[i][2]
    ffit = poly.polyval(znew, coeff) 
    
    axes[i].plot(ffit, znew)
    axes[i].plot(SSP_Input.iloc[:,i], z)
    axes[i].set_title("{}. {}".format(i, ssp))
